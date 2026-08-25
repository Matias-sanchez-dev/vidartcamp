from fastapi import FastAPI, Depends, HTTPException, Request, Form, UploadFile, File, status
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import datetime, timedelta, date, time as dt_time
from typing import Optional, List
import uuid
import urllib.parse
import io
import json
from sqlalchemy import desc, func
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

from database import engine, get_db, Base
from models import Usuario, Torneo, Equipo, Jugador, Partido, SesionQR, Gol

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI(title="Vidartcamp Control de Acceso v2")

# Static files (for images, CSS, etc.)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Templates
templates = Jinja2Templates(directory="templates")

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# JWT settings
SECRET_KEY = "vidartcamp-secret-key-change-in-production"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480
QR_EXPIRATION_SECONDS = 120

PORTERO_USERNAME = "portero"


def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password):
    return pwd_context.hash(password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def get_current_user(request: Request, db: Session = Depends(get_db)):
    """Get current logged user (jugador)"""
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        dni: str = payload.get("sub")
        user_type: str = payload.get("type", "jugador")
        if dni is None:
            return None
    except JWTError:
        return None
    
    if user_type == "jugador":
        jugador = db.query(Jugador).filter(Jugador.dni == dni).first()
        return jugador
    return None


def get_current_admin(request: Request, db: Session = Depends(get_db)):
    """Get current admin user"""
    token = request.cookies.get("admin_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        user_type: str = payload.get("type")
        if username is None or user_type != "admin":
            return None
    except JWTError:
        return None
    
    usuario = db.query(Usuario).filter(
        Usuario.username == username,
        Usuario.activo == True,
        Usuario.es_admin == True
    ).first()
    if not usuario:
        return None
    return usuario


def get_current_staff(request: Request, db: Session = Depends(get_db)):
    """Get current staff user (admin or portero)"""
    token = request.cookies.get("admin_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        user_type: str = payload.get("type")
        if username is None or user_type not in {"admin", "portero"}:
            return None
    except JWTError:
        return None

    if user_type == "admin":
        return db.query(Usuario).filter(
            Usuario.username == username,
            Usuario.activo == True,
            Usuario.es_admin == True
        ).first()

    return db.query(Usuario).filter(
        Usuario.username == username,
        Usuario.activo == True,
        Usuario.es_admin == False,
        Usuario.username == PORTERO_USERNAME
    ).first()


@app.on_event("startup")
async def startup_event():
    """Initialize database with test data"""
    db = next(get_db())
    
    # Create default admin if doesn't exist
    if not db.query(Usuario).filter(Usuario.username == "admin").first():
        admin = Usuario(
            username="admin",
            email="admin@vidartcamp.com",
            password_hash=get_password_hash("admin123"),
            es_admin=True,
            activo=True,
            fecha_creacion=datetime.utcnow()
        )
        db.add(admin)
        db.commit()
        print("✅ Usuario admin creado: username='admin', password='admin123'")

    # Create default portero if doesn't exist
    if not db.query(Usuario).filter(Usuario.username == PORTERO_USERNAME).first():
        portero = Usuario(
            username=PORTERO_USERNAME,
            email="portero@vidartcamp.com",
            password_hash=get_password_hash("portero123"),
            es_admin=False,
            activo=True,
            fecha_creacion=datetime.utcnow()
        )
        db.add(portero)
        db.commit()
        print("✅ Usuario portero creado: username='portero', password='portero123'")
    
    # Create sample tournament if doesn't exist
    if db.query(Torneo).count() == 0:
        torneo = Torneo(
            nombre="Torneo Apertura 2025",
            descripcion="Torneo de fútbol 7",
            fecha_inicio=date.today(),
            activo=True,
            mostrar_publico=True
        )
        db.add(torneo)
        db.commit()
        db.refresh(torneo)
        
        # Create teams
        equipo1 = Equipo(nombre="Los Rayos", torneo_id=torneo.id, activo=True)
        equipo2 = Equipo(nombre="Los Truenos", torneo_id=torneo.id, activo=True)
        equipo_externo = Equipo(nombre="Equipo Externo", torneo_id=None, activo=True)
        db.add_all([equipo1, equipo2, equipo_externo])
        db.commit()
        db.refresh(equipo1)
        db.refresh(equipo2)
        db.refresh(equipo_externo)
        
        # Create players
        jugador1 = Jugador(
            dni="123",
            nombre_completo="Juan Pérez",
            password_hash=get_password_hash("1234"),
            equipo_id=equipo1.id,
            activo=True
        )
        jugador2 = Jugador(
            dni="456",
            nombre_completo="Carlos González",
            password_hash=get_password_hash("1234"),
            equipo_id=equipo2.id,
            activo=True
        )
        jugador3 = Jugador(
            dni="789",
            nombre_completo="Pedro Martínez",
            password_hash=get_password_hash("1234"),
            equipo_id=equipo_externo.id,
            activo=True
        )
        db.add_all([jugador1, jugador2, jugador3])
        db.commit()
        
        # Create match for today
        partido = Partido(
            torneo_id=torneo.id,
            equipo_local_id=equipo1.id,
            equipo_visitante_id=equipo2.id,
            fecha=date.today(),
            hora=dt_time(18, 0),
            jornada=1,
            cancha="Cancha 1"
        )
        db.add(partido)
        db.commit()
        
        print("✅ Base de datos inicializada con datos de prueba")
    
    # Iniciar scheduler para reset diario
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        func=reset_ya_ingreso_diario,
        trigger=CronTrigger(hour=3, minute=0),  # Todos los días a las 00:00
        id='reset_daily_access',
        name='Reset diario de ya_ingreso',
        replace_existing=True
    )
    scheduler.start()
    print("✅ Scheduler iniciado - Reset diario configurado para las 00:00")
    
    db.close()


def reset_ya_ingreso_diario():
    """Resetea el estado ya_ingreso de todos los jugadores diariamente"""
    db = next(get_db())
    try:
        cantidad = db.query(Jugador).update({Jugador.ya_ingreso: False})
        db.commit()
        print(f"✅ Reset diario ejecutado: {cantidad} jugadores reseteados a ya_ingreso=False")
    except Exception as e:
        db.rollback()
        print(f"❌ Error en reset diario: {e}")
    finally:
        db.close()


# ==================== RUTAS PÚBLICAS ====================

@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db)):
    """Home page - shows active tournament info"""
    torneo_activo = db.query(Torneo).filter(Torneo.activo == True, Torneo.mostrar_publico == True).first()
    
    return templates.TemplateResponse("home.html", {
        "request": request,
        "torneo": torneo_activo
    })


@app.get("/tabla", response_class=HTMLResponse)
async def tabla_posiciones(request: Request, db: Session = Depends(get_db)):
    """Standings table - supports multiple active tournaments via ?torneo_id="""
    torneos_publicos = db.query(Torneo).filter(
        Torneo.activo == True, Torneo.mostrar_publico == True
    ).order_by(Torneo.id.desc()).all()

    torneo_id_raw = request.query_params.get("torneo_id")
    torneo = None
    if torneo_id_raw:
        try:
            torneo = db.query(Torneo).filter(
                Torneo.id == int(torneo_id_raw),
                Torneo.activo == True,
                Torneo.mostrar_publico == True
            ).first()
        except ValueError:
            pass

    if not torneo and torneos_publicos:
        torneo = torneos_publicos[0]

    if not torneo:
        jugador = get_current_user(request, db)
        if jugador and jugador.equipo and jugador.equipo.torneo_id:
            torneo = db.query(Torneo).filter(Torneo.id == jugador.equipo.torneo_id).first()
            torneos_publicos = [torneo] if torneo else []

    if not torneo:
        return templates.TemplateResponse("error.html", {
            "request": request,
            "mensaje": "No hay torneo activo en este momento"
        })

    equipos = db.query(Equipo).filter(
        Equipo.torneo_id == torneo.id,
        Equipo.activo == True
    ).order_by(
        Equipo.puntos.desc(),
        desc(Equipo.goles_favor - Equipo.goles_contra),
        Equipo.goles_favor.desc()
    ).all()

    return templates.TemplateResponse("tabla.html", {
        "request": request,
        "torneo": torneo,
        "equipos": equipos,
        "torneos_publicos": torneos_publicos,
    })


@app.get("/fixture", response_class=HTMLResponse)
async def fixture(request: Request, db: Session = Depends(get_db)):
    """Fixture - supports multiple active tournaments via ?torneo_id="""
    torneos_publicos = db.query(Torneo).filter(
        Torneo.activo == True, Torneo.mostrar_publico == True
    ).order_by(Torneo.id.desc()).all()

    torneo_id_raw = request.query_params.get("torneo_id")
    torneo = None
    if torneo_id_raw:
        try:
            torneo = db.query(Torneo).filter(
                Torneo.id == int(torneo_id_raw),
                Torneo.activo == True,
                Torneo.mostrar_publico == True
            ).first()
        except ValueError:
            pass

    if not torneo and torneos_publicos:
        torneo = torneos_publicos[0]

    if not torneo:
        jugador = get_current_user(request, db)
        if jugador and jugador.equipo and jugador.equipo.torneo_id:
            torneo = db.query(Torneo).filter(Torneo.id == jugador.equipo.torneo_id).first()
            torneos_publicos = [torneo] if torneo else []

    if not torneo:
        return templates.TemplateResponse("error.html", {
            "request": request,
            "mensaje": "No hay torneo activo en este momento"
        })

    partidos = db.query(Partido).filter(
        Partido.torneo_id == torneo.id
    ).order_by(Partido.jornada, Partido.fecha, Partido.hora).all()

    partidos_por_jornada = {}
    for partido in partidos:
        jornada = partido.jornada or 0
        if jornada not in partidos_por_jornada:
            partidos_por_jornada[jornada] = []
        partidos_por_jornada[jornada].append(partido)

    return templates.TemplateResponse("fixture.html", {
        "request": request,
        "torneo": torneo,
        "partidos_por_jornada": partidos_por_jornada,
        "torneos_publicos": torneos_publicos,
    })


@app.get("/goleadores", response_class=HTMLResponse)
async def goleadores(request: Request, db: Session = Depends(get_db)):
    """Top scorers table - supports multiple active tournaments via ?torneo_id="""
    torneos_publicos = db.query(Torneo).filter(
        Torneo.activo == True, Torneo.mostrar_publico == True
    ).order_by(Torneo.id.desc()).all()

    torneo_id_raw = request.query_params.get("torneo_id")
    torneo = None
    if torneo_id_raw:
        try:
            torneo = db.query(Torneo).filter(
                Torneo.id == int(torneo_id_raw),
                Torneo.activo == True,
                Torneo.mostrar_publico == True
            ).first()
        except ValueError:
            pass

    if not torneo and torneos_publicos:
        torneo = torneos_publicos[0]

    if not torneo:
        return templates.TemplateResponse("error.html", {
            "request": request,
            "mensaje": "No hay torneo activo en este momento"
        })

    goleadores_rows = db.query(
        Jugador.nombre_completo.label("jugador"),
        Equipo.nombre.label("equipo"),
        func.sum(Gol.cantidad).label("goles")
    ).join(Jugador, Gol.jugador_id == Jugador.id
    ).join(Partido, Gol.partido_id == Partido.id
    ).join(Equipo, Gol.equipo_id == Equipo.id
    ).filter(
        Partido.torneo_id == torneo.id
    ).group_by(Jugador.id
    ).order_by(func.sum(Gol.cantidad).desc(), Jugador.nombre_completo.asc()).all()

    return templates.TemplateResponse("goleadores.html", {
        "request": request,
        "torneo": torneo,
        "goleadores": goleadores_rows,
        "torneos_publicos": torneos_publicos,
    })


# ==================== JUGADOR LOGIN Y DASHBOARD ====================

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
async def login(
    request: Request,
    dni: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    jugador = db.query(Jugador).filter(Jugador.dni == dni, Jugador.activo == True).first()
    
    if not jugador or not verify_password(password, jugador.password_hash):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "DNI o contraseña incorrectos"
        })
    
    access_token = create_access_token(data={"sub": jugador.dni, "type": "jugador"})
    response = RedirectResponse(url="/dashboard", status_code=302)
    response.set_cookie(key="access_token", value=access_token, httponly=True, max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60)
    return response


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    jugador = get_current_user(request, db)
    if not jugador:
        return RedirectResponse(url="/login", status_code=302)
    
    # Check if player has match today
    hoy = date.today()
    tiene_partido = db.query(Partido).filter(
        Partido.fecha == hoy,
        ((Partido.equipo_local_id == jugador.equipo_id) | (Partido.equipo_visitante_id == jugador.equipo_id))
    ).first()
    
    partido_info = None
    if tiene_partido:
        oponente_id = tiene_partido.equipo_visitante_id if tiene_partido.equipo_local_id == jugador.equipo_id else tiene_partido.equipo_local_id
        oponente = db.query(Equipo).filter(Equipo.id == oponente_id).first()
        partido_info = {
            "hora": tiene_partido.hora.strftime("%H:%M"),
            "oponente": oponente.nombre if oponente else "Desconocido",
            "cancha": tiene_partido.cancha or "Por confirmar"
        }
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "jugador": jugador,
        "tiene_partido_hoy": tiene_partido is not None,
        "partido_info": partido_info
    })


@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=302)
    # Limpiar AMBAS cookies por seguridad
    response.delete_cookie("access_token")
    response.delete_cookie("admin_token")  # Por si acaso
    return response


@app.get("/cambiar-password", response_class=HTMLResponse)
async def cambiar_password_page(request: Request, db: Session = Depends(get_db)):
    """Página para cambiar contraseña (solo jugadores logueados)"""
    jugador = get_current_user(request, db)
    if not jugador:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("cambiar_password.html", {
        "request": request,
        "jugador": jugador
    })


@app.post("/cambiar-password")
async def cambiar_password(
    request: Request,
    password_actual: str = Form(...),
    password_nueva: str = Form(...),
    password_confirmacion: str = Form(...),
    db: Session = Depends(get_db)
):
    """Procesar cambio de contraseña"""
    jugador = get_current_user(request, db)
    if not jugador:
        return RedirectResponse(url="/login", status_code=302)
    
    # Validar contraseña actual
    if not verify_password(password_actual, jugador.password_hash):
        return templates.TemplateResponse("cambiar_password.html", {
            "request": request,
            "jugador": jugador,
            "error": "La contraseña actual es incorrecta"
        })
    
    # Validar que la nueva contraseña no esté vacía
    if not password_nueva or not password_nueva.strip():
        return templates.TemplateResponse("cambiar_password.html", {
            "request": request,
            "jugador": jugador,
            "error": "La nueva contraseña no puede estar vacía"
        })
    
    # Validar que las contraseñas coincidan
    if password_nueva != password_confirmacion:
        return templates.TemplateResponse("cambiar_password.html", {
            "request": request,
            "jugador": jugador,
            "error": "Las contraseñas nuevas no coinciden"
        })
    
    # Actualizar contraseña (hasheada)
    jugador.password_hash = get_password_hash(password_nueva)
    db.commit()
    
    # Redirigir al dashboard con mensaje de éxito
    return templates.TemplateResponse("cambiar_password.html", {
        "request": request,
        "jugador": jugador,
        "ok": "✅ Contraseña cambiada exitosamente"
    })


# ==================== QR GENERATION AND VALIDATION ====================

@app.post("/api/generar-qr")
async def generar_qr(request: Request, db: Session = Depends(get_db)):
    jugador = get_current_user(request, db)
    if not jugador:
        raise HTTPException(status_code=401, detail="No autorizado")
    
    token = str(uuid.uuid4())
    expiracion = datetime.utcnow() + timedelta(seconds=QR_EXPIRATION_SECONDS)
    
    sesion = SesionQR(
        token=token,
        jugador_id=jugador.id,
        fecha_expiracion=expiracion,
        usado=False
    )
    db.add(sesion)
    db.commit()
    
    return JSONResponse(content={
        "token": token,
        "expiracion_segundos": QR_EXPIRATION_SECONDS
    })


@app.get("/scanner", response_class=HTMLResponse)
async def scanner_page(request: Request, db: Session = Depends(get_db)):
    staff = get_current_staff(request, db)
    if not staff:
        return RedirectResponse(url="/admin/login", status_code=302)
    return templates.TemplateResponse("scanner.html", {"request": request})


@app.post("/api/validar-qr")
async def validar_qr(request: Request, db: Session = Depends(get_db)):
    staff = get_current_staff(request, db)
    if not staff:
        raise HTTPException(status_code=401, detail="No autorizado")

    data = await request.json()
    token = data.get("token")
    
    if not token:
        return JSONResponse(content={"status": "error", "mensaje": "Token no proporcionado"}, status_code=400)
    
    sesion = db.query(SesionQR).filter(SesionQR.token == token).first()
    
    if not sesion:
        return JSONResponse(content={"status": "error", "mensaje": "QR no válido"})
    
    # PROTECCIÓN CONTRA DOBLE ACCESO: Verificar si el jugador ya ingresó
    jugador = db.query(Jugador).filter(Jugador.id == sesion.jugador_id).first()
    
    if jugador.ya_ingreso:
        return JSONResponse(content={
            "status": "error", 
            "mensaje": "⚠️ ESTE JUGADOR YA INGRESÓ AL PREDIO (Intento de doble acceso)"
        })
    
    if sesion.usado:
        return JSONResponse(content={"status": "error", "mensaje": "QR ya utilizado"})
    
    if datetime.utcnow() > sesion.fecha_expiracion:
        return JSONResponse(content={"status": "error", "mensaje": "QR expirado"})
    
    equipo = db.query(Equipo).filter(Equipo.id == jugador.equipo_id).first()
    
    # Marcar token como usado y jugador como ingresado
    sesion.usado = True
    sesion.fecha_uso = datetime.utcnow()
    jugador.ya_ingreso = True  # Prevenir doble acceso
    db.commit()
    
    return JSONResponse(content={
        "status": "ok",
        "nombre": jugador.nombre_completo,
        "equipo": equipo.nombre,
        "dni": jugador.dni
    })


# ==================== ADMIN ROUTES ====================

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request, db: Session = Depends(get_db)):
    # Si ya tiene sesión de admin válida, redirigir al panel
    admin = get_current_admin(request, db)
    if admin:
        return RedirectResponse(url="/admin", status_code=302)
    
    # Si tiene sesión de jugador, ignorarla y mostrar login
    # (no importa que tenga access_token, debe logearse como admin)
    return templates.TemplateResponse("admin_login.html", {"request": request})


@app.post("/admin/login")
async def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    usuario = db.query(Usuario).filter(
        Usuario.username == username,
        Usuario.activo == True
    ).first()
    
    if not usuario or not verify_password(password, usuario.password_hash):
        return templates.TemplateResponse("admin_login.html", {
            "request": request,
            "error": "Usuario o contraseña incorrectos"
        })

    if usuario.es_admin == True:
        access_token = create_access_token(data={"sub": usuario.username, "type": "admin"})
        response = RedirectResponse(url="/admin", status_code=302)
        response.set_cookie(key="admin_token", value=access_token, httponly=True, max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60)
        return response

    if usuario.username == PORTERO_USERNAME:
        access_token = create_access_token(data={"sub": usuario.username, "type": "portero"})
        response = RedirectResponse(url="/scanner", status_code=302)
        response.set_cookie(key="admin_token", value=access_token, httponly=True, max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60)
        return response

    return templates.TemplateResponse("admin_login.html", {
        "request": request,
        "error": "No autorizado"
    })


@app.get("/admin", response_class=HTMLResponse)
async def admin_panel(request: Request, db: Session = Depends(get_db)):
    admin = get_current_admin(request, db)
    if not admin:
        response = RedirectResponse(url="/admin/login", status_code=302)
        # Limpiar admin_token si existe
        response.delete_cookie("admin_token")
        return response

    hoy = date.today()
    total_jugadores = db.query(Jugador).filter(Jugador.activo == True).count()
    partidos_hoy = db.query(Partido).filter(Partido.fecha == hoy).count()
    total_equipos = db.query(Equipo).filter(Equipo.activo == True).count()

    torneos = db.query(Torneo).order_by(Torneo.id.desc()).all()
    
    # Calcular estadísticas para cada torneo
    torneos_con_stats = []
    for t in torneos:
        equipos_count = db.query(Equipo).filter(Equipo.torneo_id == t.id).count()
        partidos_count = db.query(Partido).filter(Partido.torneo_id == t.id).count()
        torneos_con_stats.append({
            "torneo": t,
            "equipos_count": equipos_count,
            "partidos_count": partidos_count
        })
    
    torneo_activo = db.query(Torneo).filter(Torneo.activo == True).first()

    torneo_id_raw = request.query_params.get("torneo_id")
    torneo_seleccionado = None
    if torneo_id_raw:
        try:
            torneo_id = int(torneo_id_raw)
            torneo_seleccionado = db.query(Torneo).filter(Torneo.id == torneo_id).first()
        except ValueError:
            torneo_seleccionado = None
    if not torneo_seleccionado:
        torneo_seleccionado = torneo_activo

    equipos = db.query(Equipo).filter(Equipo.activo == True).order_by(Equipo.nombre.asc()).all()
    equipos_torneo = []
    if torneo_seleccionado:
        equipos_torneo = db.query(Equipo).filter(
            Equipo.activo == True,
            Equipo.torneo_id == torneo_seleccionado.id
        ).order_by(Equipo.nombre.asc()).all()

    partidos_de_hoy = db.query(Partido).filter(Partido.fecha == hoy).order_by(Partido.hora.asc()).all()
    partidos_torneo = []
    if torneo_seleccionado:
        partidos_torneo = db.query(Partido).filter(
            Partido.torneo_id == torneo_seleccionado.id
        ).order_by(Partido.jornada, Partido.fecha, Partido.hora).all()

    equipos_con_jugadores = []
    for equipo in equipos:
        count = db.query(Jugador).filter(
            Jugador.equipo_id == equipo.id,
            Jugador.activo == True
        ).count()
        equipos_con_jugadores.append({"equipo": equipo, "count": count})

    active_tab = request.query_params.get("tab", "dashboard")
    if active_tab not in {"dashboard", "equipos", "jugadores", "partidos", "torneos", "carga_masiva"}:
        active_tab = "dashboard"

    ok_raw = request.query_params.get("ok")
    error_raw = request.query_params.get("error")
    ok = urllib.parse.unquote(ok_raw) if ok_raw else None
    error = urllib.parse.unquote(error_raw) if error_raw else None

    return templates.TemplateResponse("admin.html", {
        "request": request,
        "admin": admin,
        "hoy": hoy,
        "torneos": torneos,
        "torneos_con_stats": torneos_con_stats,
        "torneo_activo": torneo_activo,
        "torneo_seleccionado": torneo_seleccionado,
        "total_jugadores": total_jugadores,
        "partidos_hoy": partidos_hoy,
        "total_equipos": total_equipos,
        "equipos": equipos,
        "equipos_torneo": equipos_torneo,
        "equipos_con_jugadores": equipos_con_jugadores,
        "partidos_de_hoy": partidos_de_hoy,
        "partidos_torneo": partidos_torneo,
        "active_tab": active_tab,
        "ok": ok,
        "error": error,
    })


@app.get("/admin/jugadores/equipo/{equipo_id}", response_class=HTMLResponse)
async def admin_jugadores_equipo_partial(equipo_id: int, request: Request, db: Session = Depends(get_db)):
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")

    equipo = db.query(Equipo).filter(Equipo.id == equipo_id, Equipo.activo == True).first()
    if not equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")

    jugadores = db.query(Jugador).filter(
        Jugador.equipo_id == equipo_id,
        Jugador.activo == True
    ).order_by(Jugador.nombre_completo.asc()).all()

    equipos = db.query(Equipo).filter(Equipo.activo == True).order_by(Equipo.nombre.asc()).all()

    return templates.TemplateResponse("admin_jugadores_partial.html", {
        "request": request,
        "equipo": equipo,
        "jugadores": jugadores,
        "equipos": equipos,
    })


@app.get("/admin/logout")
async def admin_logout():
    response = RedirectResponse(url="/admin/login", status_code=302)
    # Limpiar AMBAS cookies por seguridad
    response.delete_cookie("admin_token")
    response.delete_cookie("access_token")  # Por si acaso
    return response


def recalculate_torneo_positions(torneo_id: int, db: Session):
    equipos_t = db.query(Equipo).filter(Equipo.torneo_id == torneo_id, Equipo.activo == True).all()
    equipos_by_id = {e.id: e for e in equipos_t}

    for e in equipos_t:
        e.partidos_jugados = 0
        e.partidos_ganados = 0
        e.partidos_empatados = 0
        e.partidos_perdidos = 0
        e.goles_favor = 0
        e.goles_contra = 0
        e.puntos = 0

    partidos_finalizados = db.query(Partido).filter(
        Partido.torneo_id == torneo_id,
        Partido.finalizado == True,
        Partido.goles_local != None,
        Partido.goles_visitante != None
    ).all()

    for p in partidos_finalizados:
        local = equipos_by_id.get(p.equipo_local_id)
        visitante = equipos_by_id.get(p.equipo_visitante_id)
        if not local and not visitante:
            continue

        gl = int(p.goles_local)
        gv = int(p.goles_visitante)

        if local:
            local.partidos_jugados += 1
            local.goles_favor += gl
            local.goles_contra += gv
            if gl > gv:
                local.partidos_ganados += 1
                local.puntos += 3
            elif gl < gv:
                local.partidos_perdidos += 1
            else:
                local.partidos_empatados += 1
                local.puntos += 1

        if visitante:
            visitante.partidos_jugados += 1
            visitante.goles_favor += gv
            visitante.goles_contra += gl
            if gv > gl:
                visitante.partidos_ganados += 1
                visitante.puntos += 3
            elif gv < gl:
                visitante.partidos_perdidos += 1
            else:
                visitante.partidos_empatados += 1
                visitante.puntos += 1


@app.post("/admin/torneos/recalcular")
async def admin_recalcular_tabla(
    request: Request,
    torneo_id: int = Form(...),
    db: Session = Depends(get_db)
):
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")

    torneo = db.query(Torneo).filter(Torneo.id == torneo_id).first()
    if not torneo:
        msg = urllib.parse.quote("Torneo no encontrado")
        return RedirectResponse(url=f"/admin?tab=torneos&error={msg}", status_code=302)

    recalculate_torneo_positions(torneo_id, db)
    db.commit()

    msg = urllib.parse.quote("Tabla de posiciones recalculada")
    return RedirectResponse(url=f"/admin?tab=torneos&torneo_id={torneo_id}&ok={msg}", status_code=302)


@app.post("/admin/torneos")
async def admin_crear_torneo(
    request: Request,
    nombre: str = Form(...),
    descripcion: Optional[str] = Form(None),
    fecha_inicio: Optional[str] = Form(None),
    fecha_fin: Optional[str] = Form(None),
    mostrar_publico: Optional[str] = Form(None),
    activar: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    nombre_limpio = (nombre or "").strip()
    if not nombre_limpio:
        msg = urllib.parse.quote("Nombre de torneo inválido")
        return RedirectResponse(url=f"/admin?tab=torneos&error={msg}", status_code=302)

    fi = None
    ff = None
    try:
        if fecha_inicio:
            fi = date.fromisoformat(fecha_inicio)
        if fecha_fin:
            ff = date.fromisoformat(fecha_fin)
    except ValueError:
        msg = urllib.parse.quote("Fecha inválida")
        return RedirectResponse(url=f"/admin?tab=torneos&error={msg}", status_code=302)

    torneo = Torneo(
        nombre=nombre_limpio,
        descripcion=(descripcion or "").strip() or None,
        fecha_inicio=fi,
        fecha_fin=ff,
        activo=False,
        mostrar_publico=True if mostrar_publico else False,
    )
    db.add(torneo)
    db.commit()
    db.refresh(torneo)

    if activar:
        torneo.activo = True
        db.commit()

    msg = urllib.parse.quote("Torneo creado")
    return RedirectResponse(url=f"/admin?tab=torneos&torneo_id={torneo.id}&ok={msg}", status_code=302)


@app.post("/admin/torneos/activar")
async def admin_activar_torneo(
    request: Request,
    torneo_id: int = Form(...),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    torneo = db.query(Torneo).filter(Torneo.id == torneo_id).first()
    if not torneo:
        msg = urllib.parse.quote("Torneo no encontrado")
        return RedirectResponse(url=f"/admin?tab=torneos&error={msg}", status_code=302)

    torneo.activo = not torneo.activo
    db.commit()

    msg = urllib.parse.quote("Torneo activado" if torneo.activo else "Torneo desactivado")
    return RedirectResponse(url=f"/admin?tab=torneos&torneo_id={torneo.id}&ok={msg}", status_code=302)


@app.post("/admin/torneos/publico")
async def admin_toggle_torneo_publico(
    request: Request,
    torneo_id: int = Form(...),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    torneo = db.query(Torneo).filter(Torneo.id == torneo_id).first()
    if not torneo:
        msg = urllib.parse.quote("Torneo no encontrado")
        return RedirectResponse(url=f"/admin?tab=torneos&error={msg}", status_code=302)

    torneo.mostrar_publico = not bool(torneo.mostrar_publico)
    db.commit()

    msg = urllib.parse.quote("Visibilidad actualizada")
    return RedirectResponse(url=f"/admin?tab=torneos&torneo_id={torneo.id}&ok={msg}", status_code=302)


@app.post("/admin/torneos/eliminar")
async def admin_eliminar_torneo(
    request: Request,
    torneo_id: int = Form(...),
    db: Session = Depends(get_db)
):
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    torneo = db.query(Torneo).filter(Torneo.id == torneo_id).first()
    if not torneo:
        msg = urllib.parse.quote("Torneo no encontrado")
        return RedirectResponse(url=f"/admin?tab=torneos&error={msg}", status_code=302)
    
    # Eliminar goles y partidos asociados
    partido_ids = db.query(Partido.id).filter(Partido.torneo_id == torneo_id)
    db.query(Gol).filter(Gol.partido_id.in_(partido_ids)).delete(synchronize_session=False)
    db.query(Partido).filter(Partido.torneo_id == torneo_id).delete()
    
    # Desasignar equipos (poner torneo_id = NULL)
    db.query(Equipo).filter(Equipo.torneo_id == torneo_id).update({Equipo.torneo_id: None})
    
    # Eliminar torneo
    nombre_torneo = torneo.nombre
    db.delete(torneo)
    db.commit()
    
    msg = urllib.parse.quote(f"Torneo '{nombre_torneo}' eliminado (incluyendo partidos asociados)")
    return RedirectResponse(url=f"/admin?tab=torneos&ok={msg}", status_code=302)


@app.post("/admin/torneos/asignar-equipos")
async def admin_asignar_equipos_torneo(
    request: Request,
    torneo_id: int = Form(...),
    equipo_ids: List[int] = Form([]),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    torneo = db.query(Torneo).filter(Torneo.id == torneo_id).first()
    if not torneo:
        msg = urllib.parse.quote("Torneo no encontrado")
        return RedirectResponse(url=f"/admin?tab=torneos&error={msg}", status_code=302)

    selected_ids = set(int(x) for x in (equipo_ids or []))

    equipos_t = db.query(Equipo).filter(Equipo.activo == True).all()
    for e in equipos_t:
        if e.id in selected_ids:
            e.torneo_id = torneo.id
        elif e.torneo_id == torneo.id:
            e.torneo_id = None

    db.commit()
    recalculate_torneo_positions(torneo.id, db)
    db.commit()

    msg = urllib.parse.quote("Equipos asignados")
    return RedirectResponse(url=f"/admin?tab=torneos&torneo_id={torneo.id}&ok={msg}", status_code=302)


@app.post("/admin/equipos")
async def admin_crear_equipo(
    request: Request,
    nombre: str = Form(...),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    nombre_limpio = (nombre or "").strip()
    if not nombre_limpio:
        msg = urllib.parse.quote("Nombre de equipo inválido")
        return RedirectResponse(url=f"/admin?tab=equipos&error={msg}", status_code=302)

    existe = db.query(Equipo).filter(Equipo.nombre == nombre_limpio).first()
    if existe:
        msg = urllib.parse.quote("Ya existe un equipo con ese nombre")
        return RedirectResponse(url=f"/admin?tab=equipos&error={msg}", status_code=302)

    torneo_activo = db.query(Torneo).filter(Torneo.activo == True).first()
    equipo = Equipo(
        nombre=nombre_limpio,
        activo=True,
        torneo_id=torneo_activo.id if torneo_activo else None
    )
    db.add(equipo)
    db.commit()

    msg = urllib.parse.quote("Equipo creado")
    return RedirectResponse(url=f"/admin?tab=equipos&ok={msg}", status_code=302)


@app.post("/admin/equipos/eliminar")
async def admin_eliminar_equipo(
    request: Request,
    equipo_id: int = Form(...),
    db: Session = Depends(get_db)
):
    """Eliminar equipo y todos sus jugadores asociados"""
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")
    
    equipo = db.query(Equipo).filter(Equipo.id == equipo_id, Equipo.activo == True).first()
    if not equipo:
        msg = urllib.parse.quote("Equipo no encontrado")
        return RedirectResponse(url=f"/admin?tab=equipos&error={msg}", status_code=302)
    
    # Contar jugadores asociados activos
    jugadores_count = db.query(Jugador).filter(
        Jugador.equipo_id == equipo_id,
        Jugador.activo == True
    ).count()
    
    # Marcar jugadores como inactivos (soft delete)
    db.query(Jugador).filter(Jugador.equipo_id == equipo_id).update({Jugador.activo: False})
    
    # Marcar equipo como inactivo (soft delete)
    nombre_equipo = equipo.nombre
    equipo.activo = False
    db.commit()
    
    msg = urllib.parse.quote(
        f"Equipo '{nombre_equipo}' eliminado ({jugadores_count} jugador(es) también eliminado(s))"
    )
    return RedirectResponse(url=f"/admin?tab=equipos&ok={msg}", status_code=302)


@app.post("/admin/jugadores")
async def admin_crear_jugador(
    request: Request,
    nombre_completo: str = Form(...),
    dni: str = Form(...),
    equipo_id: int = Form(...),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    nombre_limpio = (nombre_completo or "").strip()
    dni_limpio = (dni or "").strip()
    if not nombre_limpio or not dni_limpio:
        msg = urllib.parse.quote("Nombre y DNI son obligatorios")
        return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    equipo = db.query(Equipo).filter(Equipo.id == equipo_id, Equipo.activo == True).first()
    if not equipo:
        msg = urllib.parse.quote("Equipo inválido")
        return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    existe_activo = db.query(Jugador).filter(Jugador.dni == dni_limpio, Jugador.activo == True).first()
    if existe_activo:
        msg = urllib.parse.quote("Ya existe un jugador activo con ese DNI")
        return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    existe_inactivo = db.query(Jugador).filter(Jugador.dni == dni_limpio, Jugador.activo == False).first()
    if existe_inactivo:
        existe_inactivo.nombre_completo = nombre_limpio
        existe_inactivo.equipo_id = equipo.id
        existe_inactivo.password_hash = get_password_hash(dni_limpio)
        existe_inactivo.activo = True
        db.commit()
        msg = urllib.parse.quote("Jugador reactivado")
        return RedirectResponse(url=f"/admin?tab=jugadores&ok={msg}", status_code=302)

    jugador = Jugador(
        dni=dni_limpio,
        nombre_completo=nombre_limpio,
        password_hash=get_password_hash(dni_limpio),
        equipo_id=equipo.id,
        activo=True
    )
    db.add(jugador)
    db.commit()

    msg = urllib.parse.quote("Jugador inscripto")
    return RedirectResponse(url=f"/admin?tab=jugadores&ok={msg}", status_code=302)


@app.post("/admin/jugadores/editar")
async def admin_editar_jugador(
    request: Request,
    jugador_id: int = Form(...),
    nombre_completo: str = Form(...),
    dni: str = Form(...),
    equipo_id: int = Form(...),
    telefono: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    jugador = db.query(Jugador).filter(Jugador.id == jugador_id, Jugador.activo == True).first()
    if not jugador:
        msg = urllib.parse.quote("Jugador no encontrado")
        return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    nombre_limpio = (nombre_completo or "").strip()
    dni_limpio = (dni or "").strip()
    telefono_limpio = (telefono or "").strip() or None
    email_limpio = (email or "").strip() or None

    if not nombre_limpio or not dni_limpio:
        msg = urllib.parse.quote("Nombre y DNI son obligatorios")
        return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    equipo = db.query(Equipo).filter(Equipo.id == equipo_id, Equipo.activo == True).first()
    if not equipo:
        msg = urllib.parse.quote("Equipo inválido")
        return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    if jugador.dni != dni_limpio:
        existe = db.query(Jugador).filter(Jugador.dni == dni_limpio, Jugador.activo == True).first()
        if existe:
            msg = urllib.parse.quote("Ya existe un jugador activo con ese DNI")
            return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    jugador.nombre_completo = nombre_limpio
    jugador.dni = dni_limpio
    jugador.telefono = telefono_limpio
    jugador.email = email_limpio
    jugador.equipo_id = equipo.id
    db.commit()

    msg = urllib.parse.quote("Jugador actualizado")
    return RedirectResponse(url=f"/admin?tab=jugadores&ok={msg}", status_code=302)


@app.post("/admin/jugadores/borrar")
async def admin_borrar_jugador(
    request: Request,
    jugador_id: int = Form(...),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    jugador = db.query(Jugador).filter(Jugador.id == jugador_id, Jugador.activo == True).first()
    if not jugador:
        msg = urllib.parse.quote("Jugador no encontrado")
        return RedirectResponse(url=f"/admin?tab=jugadores&error={msg}", status_code=302)

    jugador.activo = False
    db.commit()

    msg = urllib.parse.quote("Jugador eliminado de la lista")
    return RedirectResponse(url=f"/admin?tab=jugadores&ok={msg}", status_code=302)


@app.post("/admin/partidos")
async def admin_crear_partido_hoy(
    request: Request,
    torneo_id: Optional[int] = Form(None),
    fecha: Optional[str] = Form(None),
    equipo_local_id: int = Form(...),
    equipo_visitante_id: int = Form(...),
    hora: str = Form(...),
    jornada: Optional[int] = Form(None),
    cancha: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    if equipo_local_id == equipo_visitante_id:
        msg = urllib.parse.quote("Local y visitante no pueden ser el mismo equipo")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    equipo_local = db.query(Equipo).filter(Equipo.id == equipo_local_id, Equipo.activo == True).first()
    equipo_visitante = db.query(Equipo).filter(Equipo.id == equipo_visitante_id, Equipo.activo == True).first()
    if not equipo_local or not equipo_visitante:
        msg = urllib.parse.quote("Equipos inválidos")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    try:
        hora_dt = dt_time.fromisoformat(hora)
    except ValueError:
        msg = urllib.parse.quote("Hora inválida (formato HH:MM)")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    torneo_activo = db.query(Torneo).filter(Torneo.activo == True).first()
    torneo = None
    if torneo_id is not None:
        torneo = db.query(Torneo).filter(Torneo.id == torneo_id).first()
    if torneo is None:
        torneo = torneo_activo
    if not torneo:
        msg = urllib.parse.quote("No hay torneo seleccionado")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    try:
        fecha_dt = date.fromisoformat(fecha) if fecha else date.today()
    except ValueError:
        msg = urllib.parse.quote("Fecha inválida")
        return RedirectResponse(url=f"/admin?tab=partidos&torneo_id={torneo.id}&error={msg}", status_code=302)

    if equipo_local.torneo_id != torneo.id or equipo_visitante.torneo_id != torneo.id:
        msg = urllib.parse.quote("Los equipos deben pertenecer al torneo seleccionado")
        return RedirectResponse(url=f"/admin?tab=partidos&torneo_id={torneo.id}&error={msg}", status_code=302)

    partido = Partido(
        torneo_id=torneo.id,
        equipo_local_id=equipo_local.id,
        equipo_visitante_id=equipo_visitante.id,
        fecha=fecha_dt,
        hora=hora_dt,
        jornada=jornada,
        cancha=(cancha or "").strip() or None,
    )
    db.add(partido)
    db.commit()

    msg = urllib.parse.quote("Partido creado")
    return RedirectResponse(url=f"/admin?tab=partidos&torneo_id={torneo.id}&ok={msg}", status_code=302)


@app.get("/api/admin/partidos/{partido_id}/detalle")
async def admin_partido_detalle(partido_id: int, request: Request, db: Session = Depends(get_db)):
    """Datos del partido + jugadores de ambos equipos + goleadores cargados (para el modal de resultado)"""
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")

    partido = db.query(Partido).filter(Partido.id == partido_id).first()
    if not partido:
        raise HTTPException(status_code=404, detail="Partido no encontrado")

    def jugadores_de(equipo_id):
        jugadores = db.query(Jugador).filter(
            Jugador.equipo_id == equipo_id,
            Jugador.activo == True
        ).order_by(Jugador.nombre_completo.asc()).all()
        return [{"id": j.id, "nombre": j.nombre_completo} for j in jugadores]

    goles = db.query(Gol).filter(Gol.partido_id == partido.id).all()

    return JSONResponse(content={
        "partido": {
            "id": partido.id,
            "goles_local": partido.goles_local,
            "goles_visitante": partido.goles_visitante,
            "finalizado": partido.finalizado,
            "local": {"id": partido.equipo_local_id, "nombre": partido.equipo_local.nombre},
            "visitante": {"id": partido.equipo_visitante_id, "nombre": partido.equipo_visitante.nombre},
        },
        "jugadores_local": jugadores_de(partido.equipo_local_id),
        "jugadores_visitante": jugadores_de(partido.equipo_visitante_id),
        "goles": [{"jugador_id": g.jugador_id, "equipo_id": g.equipo_id, "cantidad": g.cantidad} for g in goles],
    })


@app.post("/admin/partidos/resultado")
async def admin_cargar_resultado(
    request: Request,
    partido_id: int = Form(...),
    goles_local: int = Form(...),
    goles_visitante: int = Form(...),
    goleadores: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    admin_token = request.cookies.get("admin_token")
    user_token = request.cookies.get("access_token")
    admin = get_current_admin(request, db)
    if not admin:
        if admin_token:
            raise HTTPException(status_code=403, detail="No autorizado")
        if user_token:
            return RedirectResponse(url="/", status_code=302)
        return RedirectResponse(url="/admin/login", status_code=302)

    partido = db.query(Partido).filter(Partido.id == partido_id).first()
    if not partido:
        msg = urllib.parse.quote("Partido no encontrado")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    partido.goles_local = int(goles_local)
    partido.goles_visitante = int(goles_visitante)
    partido.finalizado = True

    # Reemplazar goleadores del partido
    db.query(Gol).filter(Gol.partido_id == partido.id).delete()
    if goleadores:
        try:
            items = json.loads(goleadores)
        except (ValueError, TypeError):
            items = []

        equipos_partido = {partido.equipo_local_id, partido.equipo_visitante_id}
        cantidades = {}
        for item in items if isinstance(items, list) else []:
            try:
                jugador_id = int(item.get("jugador_id"))
                cantidad = int(item.get("cantidad", 1))
            except (ValueError, TypeError, AttributeError):
                continue
            if cantidad < 1:
                continue
            cantidades[jugador_id] = cantidades.get(jugador_id, 0) + cantidad

        for jugador_id, cantidad in cantidades.items():
            jugador = db.query(Jugador).filter(
                Jugador.id == jugador_id,
                Jugador.activo == True,
                Jugador.equipo_id.in_(equipos_partido)
            ).first()
            if not jugador:
                continue
            db.add(Gol(
                partido_id=partido.id,
                jugador_id=jugador.id,
                equipo_id=jugador.equipo_id,
                cantidad=min(cantidad, 99),
            ))

    db.commit()
    db.refresh(partido)

    if partido.torneo_id:
        recalculate_torneo_positions(partido.torneo_id, db)
        db.commit()

    msg = urllib.parse.quote("Resultado guardado")
    tid = partido.torneo_id
    torneo_qs = f"&torneo_id={tid}" if tid else ""
    return RedirectResponse(url=f"/admin?tab=partidos{torneo_qs}&ok={msg}", status_code=302)


@app.post("/admin/partidos/editar")
async def admin_editar_partido(
    request: Request,
    partido_id: int = Form(...),
    fecha: str = Form(...),
    hora: str = Form(...),
    jornada: Optional[int] = Form(None),
    cancha: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")

    partido = db.query(Partido).filter(Partido.id == partido_id).first()
    if not partido:
        msg = urllib.parse.quote("Partido no encontrado")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    torneo_qs = f"&torneo_id={partido.torneo_id}" if partido.torneo_id else ""

    try:
        partido.fecha = date.fromisoformat(fecha)
        partido.hora = dt_time.fromisoformat(hora)
    except ValueError:
        msg = urllib.parse.quote("Fecha u hora inválida")
        return RedirectResponse(url=f"/admin?tab=partidos{torneo_qs}&error={msg}", status_code=302)

    partido.jornada = jornada
    partido.cancha = (cancha or "").strip() or None
    db.commit()

    msg = urllib.parse.quote("Partido actualizado")
    return RedirectResponse(url=f"/admin?tab=partidos{torneo_qs}&ok={msg}", status_code=302)


@app.post("/admin/partidos/eliminar")
async def admin_eliminar_partido(
    request: Request,
    partido_id: int = Form(...),
    db: Session = Depends(get_db)
):
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")

    partido = db.query(Partido).filter(Partido.id == partido_id).first()
    if not partido:
        msg = urllib.parse.quote("Partido no encontrado")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    torneo_id = partido.torneo_id
    era_finalizado = partido.finalizado

    db.query(Gol).filter(Gol.partido_id == partido.id).delete()
    db.delete(partido)
    db.commit()

    if torneo_id and era_finalizado:
        recalculate_torneo_positions(torneo_id, db)
        db.commit()

    msg = urllib.parse.quote("Partido eliminado")
    torneo_qs = f"&torneo_id={torneo_id}" if torneo_id else ""
    return RedirectResponse(url=f"/admin?tab=partidos{torneo_qs}&ok={msg}", status_code=302)


@app.post("/admin/fixture/generar")
async def admin_generar_fixture(
    request: Request,
    torneo_id: int = Form(...),
    fecha_inicio: str = Form(...),
    hora: str = Form("18:00"),
    intervalo_dias: int = Form(7),
    ida_vuelta: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """Genera fixture todos-contra-todos (método del círculo) para los equipos del torneo"""
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")

    torneo = db.query(Torneo).filter(Torneo.id == torneo_id).first()
    if not torneo:
        msg = urllib.parse.quote("Torneo no encontrado")
        return RedirectResponse(url=f"/admin?tab=partidos&error={msg}", status_code=302)

    partidos_existentes = db.query(Partido).filter(Partido.torneo_id == torneo_id).count()
    if partidos_existentes > 0:
        msg = urllib.parse.quote(
            "El torneo ya tiene partidos cargados. Eliminá los partidos existentes antes de generar el fixture automático."
        )
        return RedirectResponse(url=f"/admin?tab=partidos&torneo_id={torneo_id}&error={msg}", status_code=302)

    equipos_t = db.query(Equipo).filter(
        Equipo.torneo_id == torneo_id,
        Equipo.activo == True
    ).order_by(Equipo.nombre.asc()).all()

    if len(equipos_t) < 2:
        msg = urllib.parse.quote("El torneo necesita al menos 2 equipos asignados para generar el fixture.")
        return RedirectResponse(url=f"/admin?tab=partidos&torneo_id={torneo_id}&error={msg}", status_code=302)

    try:
        fecha_base = date.fromisoformat(fecha_inicio)
        hora_dt = dt_time.fromisoformat(hora)
        intervalo = max(1, int(intervalo_dias))
    except ValueError:
        msg = urllib.parse.quote("Fecha, hora o intervalo inválido")
        return RedirectResponse(url=f"/admin?tab=partidos&torneo_id={torneo_id}&error={msg}", status_code=302)

    # Método del círculo (round-robin)
    ids = [e.id for e in equipos_t]
    if len(ids) % 2 == 1:
        ids.append(None)  # fecha libre
    n = len(ids)

    jornadas = []
    rotacion = ids[:]
    for ronda in range(n - 1):
        pares = []
        for i in range(n // 2):
            a, b = rotacion[i], rotacion[n - 1 - i]
            if a is None or b is None:
                continue
            # Alternar localía por ronda para repartir partidos de local
            pares.append((a, b) if ronda % 2 == 0 else (b, a))
        jornadas.append(pares)
        rotacion = [rotacion[0]] + [rotacion[-1]] + rotacion[1:-1]

    if ida_vuelta:
        jornadas += [[(b, a) for (a, b) in pares] for pares in jornadas]

    creados = 0
    for numero_jornada, pares in enumerate(jornadas, 1):
        fecha_jornada = fecha_base + timedelta(days=(numero_jornada - 1) * intervalo)
        for local_id, visitante_id in pares:
            db.add(Partido(
                torneo_id=torneo_id,
                equipo_local_id=local_id,
                equipo_visitante_id=visitante_id,
                fecha=fecha_jornada,
                hora=hora_dt,
                jornada=numero_jornada,
            ))
            creados += 1
    db.commit()

    msg = urllib.parse.quote(
        f"Fixture generado: {len(jornadas)} jornada(s), {creados} partido(s). Ajustá horarios y canchas desde la lista."
    )
    return RedirectResponse(url=f"/admin?tab=partidos&torneo_id={torneo_id}&ok={msg}", status_code=302)


@app.get("/admin/equipos/{equipo_id}/planilla")
async def admin_descargar_planilla_equipo(equipo_id: int, request: Request, db: Session = Depends(get_db)):
    """Planilla imprimible del equipo (Excel) para firmar con DNI el día del partido"""
    admin = get_current_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=302)

    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()
    if not equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")

    jugadores = db.query(Jugador).filter(
        Jugador.equipo_id == equipo_id,
        Jugador.activo == True
    ).order_by(Jugador.nombre_completo.asc()).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Título
    ws.merge_cells("A1:D1")
    cell = ws.cell(row=1, column=1, value=f"PLANILLA DE JUGADORES - {equipo.nombre.upper()}")
    cell.font = Font(name="Arial", bold=True, size=14)
    cell.alignment = center
    ws.row_dimensions[1].height = 26

    subtitulo = equipo.torneo.nombre if equipo.torneo else "Sin torneo asignado"
    ws.merge_cells("A2:D2")
    cell = ws.cell(row=2, column=1, value=subtitulo)
    cell.font = Font(name="Arial", size=11, color="475569")
    cell.alignment = center

    ws.merge_cells("A3:D3")
    cell = ws.cell(row=3, column=1, value="Fecha: ______________    Rival: ____________________    Jornada: ______")
    cell.font = Font(name="Arial", size=10)
    ws.row_dimensions[3].height = 20

    # Encabezados
    fila_header = 5
    headers = ["N°", "Apellido y Nombre", "DNI", "Firma"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=fila_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[fila_header].height = 22

    fila = fila_header + 1
    for i, jugador in enumerate(jugadores, 1):
        ws.cell(row=fila, column=1, value=i).alignment = center
        ws.cell(row=fila, column=2, value=jugador.nombre_completo)
        ws.cell(row=fila, column=3, value=jugador.dni).alignment = center
        for col in range(1, 5):
            c = ws.cell(row=fila, column=col)
            c.border = border
            c.font = Font(name="Arial", size=11)
        ws.row_dimensions[fila].height = 26  # espacio para la firma
        fila += 1

    # Filas en blanco extra por si hay que anotar a mano
    for _ in range(3):
        for col in range(1, 5):
            ws.cell(row=fila, column=col).border = border
        ws.row_dimensions[fila].height = 26
        fila += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30

    # Configuración de impresión
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = urllib.parse.quote(f"planilla_{equipo.nombre.replace(' ', '_').lower()}.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{nombre_archivo}"},
    )


@app.get("/admin/carga-masiva/plantilla")
async def admin_descargar_plantilla(request: Request, db: Session = Depends(get_db)):
    """Descargar plantilla Excel de ejemplo para carga masiva"""
    admin = get_current_admin(request, db)
    if not admin:
        return RedirectResponse(url="/admin/login", status_code=302)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jugadores"

    # Estilos
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Encabezados
    headers = ["Nombre", "Apellido", "DNI"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Filas de ejemplo
    ejemplos = [
        ("Juan", "Pérez", "12345678"),
        ("María", "González", "87654321"),
        ("Carlos", "López", "11223344"),
    ]
    example_font = Font(name="Arial", color="999999", italic=True, size=10)
    for row, (nombre, apellido, dni) in enumerate(ejemplos, 2):
        ws.cell(row=row, column=1, value=nombre).font = example_font
        ws.cell(row=row, column=2, value=apellido).font = example_font
        ws.cell(row=row, column=3, value=dni).font = example_font
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border

    # Ancho de columnas
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18

    # Hoja de instrucciones
    ws_info = wb.create_sheet("Instrucciones")
    instrucciones = [
        "INSTRUCCIONES PARA CARGA MASIVA DE JUGADORES",
        "",
        "1. Completá la hoja 'Jugadores' con los datos de cada jugador.",
        "2. Las columnas obligatorias son: Nombre, Apellido y DNI.",
        "3. El DNI debe ser único para cada jugador.",
        "4. Borrá las filas de ejemplo antes de cargar tus datos.",
        "5. Al subir el archivo, seleccioná el equipo al que pertenecen los jugadores.",
        "6. Si el equipo no existe, se creará automáticamente con el nombre que indiques.",
        "7. La contraseña inicial de cada jugador será su DNI.",
        "",
        "IMPORTANTE: No modifiques los nombres de las columnas (Nombre, Apellido, DNI).",
    ]
    for i, line in enumerate(instrucciones, 1):
        cell = ws_info.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(name="Arial", bold=True, size=13, color="2563EB")
        else:
            cell.font = Font(name="Arial", size=11)
    ws_info.column_dimensions["A"].width = 70

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_jugadores_vidartcamp.xlsx"},
    )


@app.post("/admin/carga-masiva")
async def admin_carga_masiva(
    request: Request,
    nombre_equipo: str = Form(""),
    equipo_id: Optional[int] = Form(None),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Carga masiva de jugadores desde un archivo Excel (.xlsx)"""
    admin = get_current_admin(request, db)
    if not admin:
        raise HTTPException(status_code=403, detail="No autorizado")

    # Validar extensión
    if not archivo.filename or not archivo.filename.lower().endswith(".xlsx"):
        msg = urllib.parse.quote("El archivo debe ser un Excel (.xlsx)")
        return RedirectResponse(url=f"/admin?tab=carga_masiva&error={msg}", status_code=302)

    # Leer archivo
    import openpyxl
    try:
        contenido = await archivo.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido))
    except Exception:
        msg = urllib.parse.quote("No se pudo leer el archivo Excel. Verificá que sea un .xlsx válido.")
        return RedirectResponse(url=f"/admin?tab=carga_masiva&error={msg}", status_code=302)

    ws = wb.active

    # Detectar columnas por nombre en la primera fila
    header_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            header_map[str(val).strip().lower()] = col

    col_nombre = header_map.get("nombre")
    col_apellido = header_map.get("apellido")
    col_dni = header_map.get("dni")

    if not col_nombre or not col_dni:
        msg = urllib.parse.quote("El archivo debe tener al menos las columnas 'Nombre' y 'DNI' en la primera fila.")
        return RedirectResponse(url=f"/admin?tab=carga_masiva&error={msg}", status_code=302)

    # Determinar equipo: usar existente o crear nuevo
    nombre_equipo_limpio = (nombre_equipo or "").strip()
    if equipo_id and equipo_id > 0:
        equipo = db.query(Equipo).filter(Equipo.id == equipo_id, Equipo.activo == True).first()
        if not equipo:
            msg = urllib.parse.quote("El equipo seleccionado no existe o fue eliminado.")
            return RedirectResponse(url=f"/admin?tab=carga_masiva&error={msg}", status_code=302)
    elif nombre_equipo_limpio:
        equipo = db.query(Equipo).filter(Equipo.nombre == nombre_equipo_limpio, Equipo.activo == True).first()
        if not equipo:
            torneo_activo = db.query(Torneo).filter(Torneo.activo == True).first()
            equipo = Equipo(
                nombre=nombre_equipo_limpio,
                activo=True,
                torneo_id=torneo_activo.id if torneo_activo else None,
            )
            db.add(equipo)
            db.commit()
            db.refresh(equipo)
    else:
        msg = urllib.parse.quote("Debés seleccionar un equipo existente o escribir un nombre para crear uno nuevo.")
        return RedirectResponse(url=f"/admin?tab=carga_masiva&error={msg}", status_code=302)

    # Procesar filas
    jugadores_creados = 0
    jugadores_duplicados = 0
    errores_fila = []

    for row in range(2, ws.max_row + 1):
        nombre_val = ws.cell(row=row, column=col_nombre).value
        apellido_val = ws.cell(row=row, column=col_apellido).value if col_apellido else ""
        dni_val = ws.cell(row=row, column=col_dni).value

        # Saltar filas vacías
        if not nombre_val and not dni_val:
            continue

        nombre_str = str(nombre_val or "").strip()
        apellido_str = str(apellido_val or "").strip()
        dni_str = str(dni_val or "").strip().replace(".0", "").replace(".", "")

        if not nombre_str or not dni_str:
            errores_fila.append(f"Fila {row}: nombre o DNI vacío")
            continue

        nombre_completo = f"{nombre_str} {apellido_str}".strip()

        # Verificar si ya existe activo
        existe_activo = db.query(Jugador).filter(Jugador.dni == dni_str, Jugador.activo == True).first()
        if existe_activo:
            jugadores_duplicados += 1
            continue

        # Si existe inactivo, reactivarlo
        existe_inactivo = db.query(Jugador).filter(Jugador.dni == dni_str, Jugador.activo == False).first()
        if existe_inactivo:
            existe_inactivo.nombre_completo = nombre_completo
            existe_inactivo.equipo_id = equipo.id
            existe_inactivo.password_hash = get_password_hash(dni_str)
            existe_inactivo.activo = True
            db.commit()
            jugadores_creados += 1
            continue

        jugador = Jugador(
            dni=dni_str,
            nombre_completo=nombre_completo,
            password_hash=get_password_hash(dni_str),
            equipo_id=equipo.id,
            activo=True,
        )
        db.add(jugador)
        jugadores_creados += 1

    db.commit()

    # Construir mensaje de resultado
    partes = [f"{jugadores_creados} jugador(es) cargado(s) al equipo '{equipo.nombre}'"]
    if jugadores_duplicados:
        partes.append(f"{jugadores_duplicados} omitido(s) por DNI duplicado")
    if errores_fila:
        partes.append(f"{len(errores_fila)} fila(s) con errores")
    msg = urllib.parse.quote(" | ".join(partes))
    return RedirectResponse(url=f"/admin?tab=carga_masiva&ok={msg}", status_code=302)


@app.get("/admin/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request, db: Session = Depends(get_db)):
    return RedirectResponse(url="/admin", status_code=302)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
